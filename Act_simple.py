import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


class Act():
    '''
    act_number  Номер Акта
    data_today  Дата создания документа
    sender  Отправитель
    carrier  Перевозчик
    recipient  Получатель

    order_info = ((number_order, comment, type_packaging, number_seats), (...), ...)
        number_order  номер заказа
        comment  комментарий к заказу
        type_packaging  тип упаковки / ед. измерения
        number_seats  количество мест
    total_size  Итого мест
    job_title  должность отправителя
    first_last_name  ФИО Отправителя
    data_out = Дата отправки
    '''

    # Создаем новую книгу и лист
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Акт"

    # Получаемые переменные
    act_number = '2323232' # Номер Акта
    data_today = "=TODAY()" # Дата создания документа
    sender = "Кострома" # Отправитель
    carrier = "СДЭК" # Перевозчик
    recipient = "Офис, г. Москва, Большой Саввинский пер, д. 12 стр. 16" # Получатель
    number_order = "55555" # номер заказа
    comment = '----' # комментарий к заказу
    type_packaging = 'короб' # тип упаковки / ед. измерения
    number_seats = 22 # количество мест
    order_info = ((number_order, comment, type_packaging, number_seats), ('8888', 'нет', 'пакет', 11),)
    total_size = sum(i[3] for i in order_info) # Итого мест
    job_title = 'Офис менеджер' # должность отправителя
    first_last_name = 'Иван Иванович Иванов' # ФИО Отправителя
    data_out = '17.02.2025' # Дата отправки

    # Настройка стилей
    header_font = Font(bold=True, size=14, name='Times New Roman') # Шрифт заголовка
    small_font = Font(bold=False, size=8, name='Times New Roman') # Шрифт заметок
    midle_font = Font(bold=False, size=11, name='Times New Roman') # Шрифт текста простого
    big_font = Font(bold=True, size=12, name='Times New Roman') # Шрифт заголовков 2-го уровня
    table_font = Font(bold=True, size=10, name='Times New Roman') # Шрифт заголовков таблицы
    header_alignment = Alignment(horizontal='center', vertical='center') # выравнивание по центру
    border_thin = Side(style='thin')
    border_thick = Side(style='thick')
    border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    border_thick_bottom = Border(bottom=border_thick)
    fill_gray = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid") # заливка серым
    aligment_left = Alignment(horizontal='left', vertical='center')
    aligment_right = Alignment(horizontal='right', vertical='center')
    aligment_left_bottom = Alignment(horizontal='left', vertical='bottom')
    aligment_right_bottom = Alignment(horizontal='right', vertical='bottom')
    aligment_center_center = Alignment(horizontal='center', vertical='center')
    aligment_wrap_true = Alignment(wrapText=True)
    font_bold_11 = Font(bold=True, size=11, name='Times New Roman')

    def template_act(self):

        # Заголовок
        self.ws.merge_cells('B1:G1') # Объединение ячеек
        self.ws['B1'] = "Акт приема-передачи груза"
        self.ws['B1'].font = self.header_font
        self.ws['B1'].alignment = self.header_alignment
        self.ws.merge_cells('C2:F2')
        self.ws['C2'] = f"№ {self.act_number}"
        self.ws['C2'].font = self.big_font
        self.ws['C2'].alignment = self.header_alignment

        # Дата
        self.ws.merge_cells('G2:H2')
        self.ws['G2'] = self.data_today
        self.ws['G2'].number_format = 'DD.MM.YYYY'
        self.ws['G2'].font = self.midle_font
        self.ws['G2'].alignment = self.aligment_right

        # Отправитель
        self.ws['A4'] = "Отправитель (название, адрес):"
        self.ws['A4'].font = self.big_font
        self.ws.merge_cells('A4:H4')
        self.ws.merge_cells('A5:H5')
        self.ws['A5'] = self.sender
        self.ws['A5'].font = self.midle_font
        self.ws['A5'].fill = self.fill_gray
        self.ws['A4'].alignment = self.aligment_left
        self.ws['A5'].alignment = self.aligment_left

        # Перевозчик
        self.ws.merge_cells('A7:H7')
        self.ws['A7'] = "Перевозчик (название, номер ТС):"
        self.ws['A7'].font = self.big_font
        self.ws['A7'].alignment = self.aligment_left
        self.ws.merge_cells('A8:H8')
        self.ws['A8'] = self.carrier
        self.ws['A8'].font = self.midle_font
        self.ws['A8'].fill = self.fill_gray
        self.ws['A8'].alignment = self.aligment_left

        # Получатель
        self.ws['A10'] = "Получатель (название, адрес):"
        self.ws['A10'].font = self.big_font
        self.ws.merge_cells('A10:H10')
        self.ws.merge_cells('A11:H11')
        self.ws['A11'] = self.recipient
        self.ws['A11'].font = self.midle_font
        self.ws['A11'].fill = self.fill_gray
        self.ws['A10'].alignment = self.aligment_left
        self.ws['A11'].alignment = self.aligment_left

        # Описание акта
        self.ws['A12'] = "Настоящий Акт свидетельствует о факте передачи Отправителем Перевозчику следующего груза для доставки Получателю:"
        self.ws.merge_cells('A12:H13')
        self.ws['A12'].font = self.font_bold_11
        self.ws['A12'].alignment = self.aligment_wrap_true # Перенос текста

        # Создание и объединение ячеек в таблице
        for row in range(15, 22):
            for col in range(1, 9):
                match col:
                    case 1:
                        self.ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
                    case 3:
                        self.ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 2)
                    case 6:
                        self.ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)


        # Заголовки таблицы
        self.ws['A15'] = "№ заказа или наименование груза"
        self.ws['C15'] = "Комментарий к грузу"
        self.ws['F15'] = "Ед. измерения/тип упаковки"
        self.ws['H15'] = "Количество мест"

        # Применение стилей к заголовкам таблицы
        for cell in self.ws['A15:H15'][0]:
            cell.font = self.table_font
            cell.border = self.border
            cell.fill = self.fill_gray
            cell.alignment = self.aligment_wrap_true

        # Пример данных
        # self.ws['A16'] = self.number_order
        # self.ws['C16'] = self.comment
        # self.ws['F16'] = self.type_packaging
        # self.ws['H16'] = self.number_seats
        cont = 0
        rows_table = self.ws['A16:H21']
        for i, row in enumerate(rows_table):
            for cel in row:
                if i < len(self.order_info):
                    if not isinstance(cel, openpyxl.cell.cell.MergedCell):
                        cel.value = self.order_info[i][cont]
                        cont += 1
            cont = 0

        # Применение стилей к данным
        for row in self.ws['A16:H21']:
            for cell in row:
                cell.font = self.midle_font
                cell.border = self.border
                cell.alignment = self.aligment_center_center

        # Итого мест
        self.ws.merge_cells('F22:G22')
        self.ws['F22'] = "Итого мест:"
        self.ws['F22'].font = self.font_bold_11
        self.ws['F22'].alignment = self.aligment_left
        self.ws['H22'] = self.total_size
        self.ws['H22'].font = self.font_bold_11
        self.ws['H22'].border = self.border
        self.ws['H22'].alignment = self.aligment_center_center

        # Подписи
        # Отправитель
        self.ws.merge_cells('A24:H24')
        self.ws['A24'] = "Со стороны Отправителя груз сдал:"
        self.ws['A24'].font = self.big_font
        self.ws.merge_cells('A25:D25')
        self.ws['A25'] = self.job_title
        self.ws['A25'].font = self.midle_font
        self.ws['A25'].alignment = self.aligment_left_bottom
        self.ws['A25'].fill = self.fill_gray
        self.ws.merge_cells('A26:B26')
        self.ws['A26'] = '(должность)'
        self.ws['A26'].alignment = self.aligment_center_center
        self.ws['A26'].font = self.small_font
        self.ws.merge_cells('E25:H25')
        self.ws['E25'] = self.first_last_name
        self.ws['E25'].font = self.midle_font
        self.ws['E25'].alignment = self.aligment_right_bottom
        self.ws['E25'].fill = self.fill_gray
        self.ws.merge_cells('C26:F26')
        self.ws['C26'] = '(подпись)'
        self.ws['C26'].alignment = self.aligment_center_center
        self.ws['C26'].font = self.small_font
        self.ws.merge_cells('G26:H26')
        self.ws['G26'] = '(фамилия, имя, отчество)'
        self.ws['G26'].alignment = self.aligment_center_center
        self.ws['G26'].font = self.small_font

        self.ws['A27'] = "Дата:"
        self.ws['A27'].font = self.midle_font
        self.ws['A27'].alignment = self.aligment_left_bottom
        self.ws['A27'].fill = self.fill_gray
        self.ws.merge_cells('B27:C27')
        self.ws['B27'] = self.data_out
        self.ws['B27'].number_format = 'DD.MM.YYYY HH:MM:SS'
        self.ws['B27'].font = self.midle_font
        self.ws['B27'].alignment = self.aligment_right_bottom
        self.ws['B27'].fill = self.fill_gray
        self.ws.merge_cells('D27:E27')
        self.ws['D27'] = 'М.П.'
        self.ws['D27'].font = self.midle_font
        self.ws['D27'].alignment = self.aligment_center_center

        # Перевозчик
        self.ws.merge_cells('A29:H30')
        self.ws['A29'] = "Груз принят Перевозчиком без внешних повреждений.\nГруз принят Перевозчиком по количеству мест, без пересчета внутренних вложений."
        self.ws['A29'].font = self.small_font
        self.ws['A29'].alignment = self.aligment_wrap_true
        self.ws.merge_cells('A31:C31')
        self.ws['A31'] = "(замечания при передаче груза)"
        self.ws['A31'].font = self.small_font

        self.ws.merge_cells('A32:H32')
        self.ws['A32'] = "Со стороны Перевозчика груз к перевозке принял:"
        self.ws['A32'].font = self.big_font
        self.ws.merge_cells('A33:D33')
        self.ws['A33'] = ''
        self.ws['A33'].font = self.midle_font
        self.ws['A33'].alignment = self.aligment_left_bottom
        self.ws['A33'].fill = self.fill_gray
        self.ws.merge_cells('A34:B34')
        self.ws['A34'] = '(должность)'
        self.ws['A34'].alignment = self.aligment_center_center
        self.ws['A34'].font = self.small_font
        self.ws.merge_cells('E33:H33')
        self.ws['E33'] = ''
        self.ws['E33'].font = self.midle_font
        self.ws['E33'].alignment = self.aligment_right_bottom
        self.ws['E33'].fill = self.fill_gray
        self.ws.merge_cells('C34:F34')
        self.ws['C34'] = '(подпись)'
        self.ws['C34'].alignment = self.aligment_center_center
        self.ws['C34'].font = self.small_font
        self.ws.merge_cells('G34:H34')
        self.ws['G34'] = '(фамилия, имя, отчество)'
        self.ws['G34'].alignment = self.aligment_center_center
        self.ws['G34'].font = self.small_font

        self.ws['A35'] = "Дата:"
        self.ws['A35'].font = self.midle_font
        self.ws['A35'].alignment = self.aligment_left_bottom
        self.ws['A35'].fill = self.fill_gray
        self.ws.merge_cells('B35:C35')
        self.ws['B35'] = ''
        self.ws['B35'].number_format = 'DD.MM.YYYY HH:MM:SS'
        self.ws['B35'].font = self.midle_font
        self.ws['B35'].alignment = self.aligment_right_bottom
        self.ws['B35'].fill = self.fill_gray
        self.ws.merge_cells('D35:E35')
        self.ws['D35'] = 'М.П.'
        self.ws['D35'].font = self.midle_font
        self.ws['D35'].alignment = self.aligment_center_center

        # Получатель
        self.ws.merge_cells('A37:H38')
        self.ws['A37'] = "Груз принят Получателем без внешних повреждений.\nГруз принят Получателем по количеству мест, без пересчета внутренних вложений."
        self.ws['A37'].font = self.small_font
        self.ws['A37'].alignment = self.aligment_wrap_true
        self.ws.merge_cells('A39:C39')
        self.ws['A39'] = "(замечания при передаче груза)"
        self.ws['A39'].font = self.small_font

        self.ws.merge_cells('A40:H40')
        self.ws['A40'] = "Со стороны Получателя груз получил:"
        self.ws['A40'].font = self.big_font
        self.ws.merge_cells('A41:D41')
        self.ws['A41'] = ''
        self.ws['A41'].font = self.midle_font
        self.ws['A41'].alignment = self.aligment_left_bottom
        self.ws['A41'].fill = self.fill_gray
        self.ws.merge_cells('A42:B42')
        self.ws['A42'] = '(должность)'
        self.ws['A42'].alignment = self.aligment_center_center
        self.ws['A42'].font = self.small_font
        self.ws.merge_cells('E41:H41')
        self.ws['E41'] = ''
        self.ws['E41'].font = self.midle_font
        self.ws['E41'].alignment = self.aligment_right_bottom
        self.ws['E41'].fill = self.fill_gray
        self.ws.merge_cells('C42:F42')
        self.ws['C42'] = '(подпись)'
        self.ws['C42'].alignment = self.aligment_center_center
        self.ws['C42'].font = self.small_font
        self.ws.merge_cells('G42:H42')
        self.ws['G42'] = '(фамилия, имя, отчество)'
        self.ws['G42'].alignment = self.aligment_center_center
        self.ws['G42'].font = self.small_font

        self.ws['A43'] = "Дата:"
        self.ws['A43'].font = self.midle_font
        self.ws['A43'].alignment = self.aligment_left_bottom
        self.ws['A43'].fill = self.fill_gray
        self.ws.merge_cells('B43:C43')
        self.ws['B43'] = ''
        self.ws['B43'].number_format = 'DD.MM.YYYY HH:MM:SS'
        self.ws['B43'].font = self.midle_font
        self.ws['B43'].alignment = self.aligment_right_bottom
        self.ws['B43'].fill = self.fill_gray
        self.ws.merge_cells('D43:E43')
        self.ws['D43'] = 'М.П.'
        self.ws['D43'].font = self.midle_font
        self.ws['D43'].alignment = self.aligment_center_center

        # Настройка ширины столбцов
        for col in range(1, 10):
            self.ws.column_dimensions[get_column_letter(col)].width = 10

        # Настройка высоты строк
        self.ws.row_dimensions[1].height = 30
        self.ws.row_dimensions[15].height = 25
        self.ws.row_dimensions[4].height = 20
        self.ws.row_dimensions[5].height = 20
        self.ws.row_dimensions[7].height = 20
        self.ws.row_dimensions[8].height = 20
        self.ws.row_dimensions[10].height = 20
        self.ws.row_dimensions[11].height = 20
        self.ws.row_dimensions[24].height = 20
        self.ws.row_dimensions[25].height = 20
        self.ws.row_dimensions[26].height = 8
        self.ws.row_dimensions[32].height = 20
        self.ws.row_dimensions[33].height = 20
        self.ws.row_dimensions[34].height = 8
        self.ws.row_dimensions[40].height = 20
        self.ws.row_dimensions[41].height = 20
        self.ws.row_dimensions[42].height = 8
        self.ws.row_dimensions[3].height = 12
        self.ws.row_dimensions[6].height = 12
        self.ws.row_dimensions[9].height = 12
        self.ws.row_dimensions[14].height = 2

        # Сохраняем Excel-файл
        excel_file = "Акт.xlsx"
        self.wb.save(excel_file)

        print(f"Файл успешно сохранен: {excel_file}")
